from django.shortcuts import render
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import random
import string
import qrcode
from io import BytesIO
from generador_qr.models import CodigoQR
import zipfile
import os

codigos_generados = set()  # Conjunto para almacenar códigos generados

def generar_codigo_aleatorio():
    longitud = 10
    caracteres = string.ascii_letters + string.digits
    while True:
        codigo = ''.join(random.choice(caracteres) for _ in range(longitud))
        if not CodigoQR.objects.filter(codigo=codigo).exists():
            break
    CodigoQR.objects.create(codigo=codigo)
    return codigo

def generar_pdf_y_excel(cantidad_categorias, categorias):
    wb = Workbook()
    ws = wb.active

    archivos_pdf = []

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
        for i, categoria_entry in enumerate(categorias.values()):
            nombre_categoria = categoria_entry[0]
            cantidad_qr = int(categoria_entry[1])

            pdf_output = BytesIO()
            pdf = canvas.Canvas(pdf_output, pagesize=A4)

            x_margin = 50
            y_margin = 50
            qr_size = 100
            qr_per_row = 4
            qr_spacing = 20

            qr_width = (A4[0] - 2 * x_margin - (qr_per_row - 1) * qr_spacing) / qr_per_row
            qr_height = qr_width

            count = 0

            pdf.setTitle(f"Codigos QR - {nombre_categoria}")

            for j in range(cantidad_qr):
                codigo = generar_codigo_aleatorio()
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=4,
                )
                qr.add_data(codigo)
                qr.make(fit=True)

                img = qr.make_image(fill_color="black", back_color="white")

                current_page = pdf.getPageNumber()
                if current_page == 0:
                    y_start = A4[1] - y_margin
                else:
                    y_start = A4[1] - y_margin - qr_height - (qr_height + qr_spacing) * ((count // qr_per_row) % 5)
                    if count % qr_per_row == 0:
                        y_start -= qr_height + qr_spacing

                pdf.drawInlineImage(img, x_margin + (qr_width + qr_spacing) * (count % qr_per_row), y_start, width=qr_width, height=qr_height)

                count += 1

                if count % (qr_per_row * 5) == 0 and j < cantidad_qr - 1:
                    pdf.showPage()  # Cambiar de página si se han generado qr_per_row códigos QR

                # Agregar el código QR generado al archivo Excel
                ws.cell(row=j + 2, column=i + 1, value=codigo)

            # Guardar el PDF en el archivo zip
            pdf.save()
            zip_file.writestr(f"Codigo_qr_{nombre_categoria}.pdf", pdf_output.getvalue())

            # Guardar los nombres de las categorías en el Excel
            ws.cell(row=1, column=i + 1, value=nombre_categoria)

        # Guardar el archivo Excel en el buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)

        # Agregar el archivo Excel al archivo ZIP
        zip_file.writestr("Codigos_QR.xlsx", excel_buffer.getvalue())

    # Devolver el archivo zip como respuesta HTTP
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="Codigos_QR.zip"'
    return response


def generar_interfaz(request):
    if request.method == 'POST':
        cantidad_categorias_str = request.POST.get('cantidad_categorias')
        if not cantidad_categorias_str:
            messages.error(request, "Por favor, ingrese la cantidad de categorías.")
            return render(request, 'generador_qr.html')

        cantidad_categorias = int(cantidad_categorias_str)

        categorias = {}  # Diccionario para almacenar las entradas de categorías y sus cantidades

        for i in range(cantidad_categorias):
            nombre_categoria = request.POST.get(f'categoria_{i + 1}')
            cantidad_qr = request.POST.get(f'cantidad_qr_{i + 1}')
            categorias[i] = (nombre_categoria, cantidad_qr)

        return generar_pdf_y_excel(cantidad_categorias, categorias)

    else:
        return render(request, 'generador_qr.html')  # Renderizar el formulario de generación de QR
